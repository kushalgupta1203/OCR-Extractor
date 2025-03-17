from flask import Flask, render_template, request, send_file, jsonify
import os
import requests
from io import BytesIO
import cloudinary
from cloudinary.uploader import upload
from cloudinary.utils import cloudinary_url
from openpyxl import Workbook
from openpyxl.styles import Font
import logging
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

app = Flask(__name__)

# Configure Cloudinary
cloudinary.config(
    cloud_name=os.getenv('CLOUDINARY_CLOUD_NAME'),
    api_key=os.getenv('CLOUDINARY_API_KEY'),
    api_secret=os.getenv('CLOUDINARY_API_SECRET')
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

processed_data_list = []
TIMEOUT = 300  # 5 minutes timeout

def upload_image_to_cloudinary(image_bytes):
    """Uploads an image to Cloudinary and returns the URL."""
    try:
        upload_result = upload(
            BytesIO(image_bytes),
            resource_type="image"
        )
        return upload_result['secure_url']
    except Exception as e:
        logger.error(f"Cloudinary upload failed: {str(e)}")
        return None

def extract_text_from_image(image_url):
    """
    Uses OCR.space to extract barcode and alphanumeric code from the image.
    """
    try:
        response = requests.post(
            "https://api.ocr.space/parse/image",
            data={
                "apikey": os.getenv("OCR_API_KEY"),
                "url": image_url,
                "language": "eng",
                "isOverlayRequired": False
            },
            timeout=TIMEOUT
        )

        if response.status_code != 200:
            logger.error(f"OCR API Error: {response.status_code} - {response.text}")
            return "OCR API Error"

        data = response.json()
        if not data.get("ParsedResults"):
            return "No text detected"

        extracted_text = data["ParsedResults"][0]["ParsedText"].strip()
        return extracted_text if extracted_text else "No valid code found"

    except requests.exceptions.Timeout:
        logger.error("OCR request timed out")
        return "OCR processing timed out"
    except Exception as e:
        logger.error(f"OCR Error: {str(e)}")
        return "OCR processing failed"

def generate_html_table():
    """Generate HTML table with error handling"""
    try:
        if not processed_data_list:
            return "<div class='alert alert-warning'>No results to display</div>"

        html = '''<table class="table table-striped table-bordered">
            <thead class="thead-dark"><tr><th>Image</th><th>Extracted Code</th></tr></thead>
            <tbody>'''
        
        for data in processed_data_list:
            html += f'''
            <tr>
                <td class="align-middle">
                    <a href="{data['Image URL']}" target="_blank" class="image-link">
                        <img src="{data['Image URL']}" class="thumbnail img-thumbnail">
                    </a>
                </td>
                <td class="align-middle">
                    <pre>{data['Extracted Code']}</pre>
                </td>
            </tr>'''
        html += '</tbody></table>'
        return html
    
    except Exception as e:
        logger.error(f"HTML generation error: {str(e)}")
        return "<div class='alert alert-danger'>Error generating results</div>"


def process_images(image_files):
    """
    Processes each image:
    1. Uploads the image to Cloudinary.
    2. Extracts barcode & alphanumeric text using OCR.space.
    """
    global processed_data_list
    processed_data_list.clear()

    for file in image_files:
        try:
            image_bytes = file.read()

            # Upload image to Cloudinary
            image_url = upload_image_to_cloudinary(image_bytes)
            if not image_url:
                processed_data_list.append({
                    "Image URL": "N/A",
                    "Extracted Code": "Cloudinary upload failed"
                })
                continue

            # Extract barcode & text using OCR.space
            extracted_code = extract_text_from_image(image_url)

            processed_data_list.append({
                "Image URL": image_url,
                "Extracted Code": extracted_code
            })

        except Exception as e:
            logger.error(f"Error processing image: {str(e)}")
            processed_data_list.append({
                "Image URL": "N/A",
                "Extracted Code": "Processing error"
            })

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        try:
            if 'images' not in request.files:
                return jsonify({"success": False, "error": "No files uploaded"})

            files = request.files.getlist('images')
            valid_files = [f for f in files if f.filename.lower().endswith(('.png', '.jpg', '.jpeg'))]

            if not valid_files:
                return jsonify({"success": False, "error": "No valid image files"})

            process_images(valid_files)

            return jsonify({
                "success": True,
                "table_html": generate_html_table(),
                "download_url": "/download"
            })

        except Exception as e:
            logger.error(f"Main processing error: {str(e)}")
            return jsonify({"success": False, "error": "System error occurred"})

    return render_template("index.html")

@app.route("/download")
def download():
    """Excel download endpoint with error handling"""
    try:
        if not processed_data_list:
            return jsonify({"error": "No data available for download"})

        wb = Workbook()
        ws = wb.active
        ws.append(["Image URL", "Extracted Code"])

        for data in processed_data_list:
            row = [data['Image URL'], data['Extracted Code']]
            ws.append(row)

            cell = ws.cell(row=ws.max_row, column=1)
            cell.hyperlink = data['Image URL']
            cell.font = Font(color="0000FF", underline="single")

        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 30

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return send_file(
            excel_buffer,
            download_name="barcode_report.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        return jsonify({"error": "Failed to generate download file"})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
